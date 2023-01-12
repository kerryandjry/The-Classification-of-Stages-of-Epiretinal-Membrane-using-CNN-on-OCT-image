import numpy as np
import matplotlib.pyplot as plt
from itertools import cycle

from openpyxl import load_workbook
from sklearn import svm, datasets
from sklearn.metrics import roc_curve, auc
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import label_binarize
from sklearn.multiclass import OneVsRestClassifier
from sklearn.metrics import roc_auc_score

wb = load_workbook()
sheet = wb['Sheet1']

y = []
for i in range(2, 319):
    cell = sheet.cell(row=i, column=7)
    cell = cell.value
    class_list = [0, 0, 0, 0]
    class_list[cell - 1] = 1
    y.append(class_list)
y = np.array(y)

n_classes = 4
# # for dl model
# for i in range(2, 319):
#     cell = sheet.cell(row=i, column=13)
#     cell = cell.value
#     log_list = []
#     num = cell.split(',')
#     for num in num:
#         log_list.append(float(num))
#     y_score.append(log_list)
colors = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd", "#e377c2", "#bcbd22", "#17becf"]
for excel_col in range(9, 17):
    y_score = []
    model_match = {9: "ResNet34", 10: "MobileNet", 11: "EfficientNet", 12: "Swin Transformer",
                   13: "Mlp-Mixer", 14: "F1", 15: "F2", 16: "F3"}
    for i in range(2, 319):
        cell = sheet.cell(row=i, column=excel_col)
        cell = cell.value
        log_list = []
        num = cell.split(',')
        for num in num:
            log_list.append(float(num))
        y_score.append(log_list)
    y_score = np.array(y_score)

    # Compute ROC curve and ROC area for each class
    fpr = dict()
    tpr = dict()
    roc_auc = dict()
    for i in range(n_classes):
        fpr[i], tpr[i], _ = roc_curve(y[:, i], y_score[:, i])
        roc_auc[i] = auc(fpr[i], tpr[i])
    #
    # Compute micro-average ROC curve and ROC area
    fpr["micro"], tpr["micro"], _ = roc_curve(y.ravel(), y_score.ravel())
    roc_auc["micro"] = auc(fpr["micro"], tpr["micro"])
    lw = 2

    # First aggregate all false positive rates
    all_fpr = np.unique(np.concatenate([fpr[i] for i in range(n_classes)]))

    # Then interpolate all ROC curves at this points
    mean_tpr = np.zeros_like(all_fpr)
    for i in range(n_classes):
        mean_tpr += np.interp(all_fpr, fpr[i], tpr[i])

    # Finally average it and compute AUC
    mean_tpr /= n_classes

    # fpr["macro"] = all_fpr
    # tpr["macro"] = mean_tpr
    # roc_auc["macro"] = auc(fpr["macro"], tpr["macro"])

    plt.plot(
        fpr["micro"],
        tpr["micro"],
        label="micro-average ROC curve of {0} (area = {1:0.2f})".format(model_match[excel_col], roc_auc["micro"]),
        color=colors[excel_col - 9],
        linewidth=2,
    )
for excel_col in range(2, 6):
    y_score = []
    colors = [0, 0, "aqua", "darkorange", "cornflowerblue", "darkolivegreen"]
    for i in range(2, 319):
        cell = sheet.cell(row=i, column=excel_col)
        cell = cell.value
        class_list = [0, 0, 0, 0]
        class_list[cell - 1] = 50
        y_score.append(class_list)
    y_score = np.array(y_score)

    # Compute ROC curve and ROC area for each class
    fpr = dict()
    tpr = dict()
    roc_auc = dict()
    for i in range(n_classes):
        fpr[i], tpr[i], _ = roc_curve(y[:, i], y_score[:, i])
        roc_auc[i] = auc(fpr[i], tpr[i])
    #
    # Compute micro-average ROC curve and ROC area
    fpr["micro"], tpr["micro"], _ = roc_curve(y.ravel(), y_score.ravel())
    roc_auc["micro"] = auc(fpr["micro"], tpr["micro"])
    lw = 2

    # First aggregate all false positive rates
    all_fpr = np.unique(np.concatenate([fpr[i] for i in range(n_classes)]))

    # Then interpolate all ROC curves at this points
    mean_tpr = np.zeros_like(all_fpr)
    for i in range(n_classes):
        mean_tpr += np.interp(all_fpr, fpr[i], tpr[i])

    # Finally average it and compute AUC
    mean_tpr /= n_classes

    fpr["macro"] = all_fpr
    tpr["macro"] = mean_tpr
    roc_auc["macro"] = auc(fpr["macro"], tpr["macro"])

    # plt.plot(
    #     fpr["micro"],
    #     tpr["micro"],
    #     # label="micro-average ROC curve of doctor {0} (area = {1:0.2f})".format(excel_col - 1, roc_auc["micro"]),
    #     color=colors[excel_col],
    #     linestyle=":",
    #     linewidth=2
    # )
x = [0.102, 0.116, 0.1599, 0.165]
y = [0.685, 0.67, 0.514, 0.5]
ave_x = [0.1357]
ave_y = [0.591]
plt.plot(x, y, 'x')
s = [20*2**1 for n in range(1)]
plt.plot(ave_x, ave_y, '+', color="red", markersize=15)

plt.plot([0, 1], [0, 1], "k--", lw=2)
plt.xlim([0.0, 1.0])
plt.ylim([0.0, 1.05])
plt.xlabel("False Positive Rate")
plt.ylabel("True Positive Rate")
plt.title("Receiver operating characteristic")
plt.legend(loc="center left", bbox_to_anchor=(0.2, 0.2), prop={"size":9})
plt.show()
plt.savefig("t.png", dpi=500)
