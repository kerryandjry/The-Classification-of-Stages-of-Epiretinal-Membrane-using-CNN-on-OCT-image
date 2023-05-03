import os
import json
from pathlib import Path
import torch
from PIL import Image
from torchvision import transforms
import matplotlib.pyplot as plt

from models.vit_model import vit_base_patch16_224_in21k
from models.swin_model import swin_tiny_patch4_window7_224
from models.resnet import resnet34
from models.efficient import efficientnetv2_s
from models.mobilenet import mobilenet_v3_small
from models import mlp

from openpyxl import load_workbook, Workbook


def main():
    wb = Workbook()
    sheet = wb.create_sheet('test')

    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")

    model_eff = efficientnetv2_s(num_classes=4).to(device)
    model_res = resnet34(num_classes=4).to(device)
    model_mob = mobilenet_v3_small(num_classes=4).to(device)
    model_vit = vit_base_patch16_224_in21k(num_classes=4).to(device)
    model_swin = swin_tiny_patch4_window7_224(num_classes=4).to(device)
    model_mlp = mlp.linear_tiny(num_classes=4).to(device)
    model_weight_path_eff = "weights/efficient/eff_4.pth"
    model_weight_path_res = "weights/resnet34/res_4.pth"
    model_weight_path_mob = "weights/mobile/mob_4.pth"
    model_weight_path_vit = "weights/vit/vit_4.pth"
    model_weight_path_swin = "weights/swin/swin_4.pth"
    model_weight_path_mlp = "weights/mlp/tiny_4.pth"


    model_eff.load_state_dict(torch.load(model_weight_path_eff, map_location=device))
    model_res.load_state_dict(torch.load(model_weight_path_res, map_location=device))
    model_mob.load_state_dict(torch.load(model_weight_path_mob, map_location=device))
    model_vit.load_state_dict(torch.load(model_weight_path_vit, map_location=device))
    model_swin.load_state_dict(torch.load(model_weight_path_swin, map_location=device))
    model_mlp.load_state_dict(torch.load(model_weight_path_mlp, map_location=device))
    model_eff.eval()
    model_res.eval()
    model_mob.eval()
    model_vit.eval()
    model_swin.eval()
    model_mlp.eval()

    model_list = [model_mob, model_eff, model_res, model_vit, model_swin, model_mlp]

    data_transform = transforms.Compose([transforms.Resize([224, 375]),
                                         transforms.CenterCrop(224),
                                         transforms.ToTensor(),
                                         transforms.Normalize([0.5, 0.5, 0.5], [0.5, 0.5, 0.5])
                                         ])
    path = Path("./test/2431")
    res_list = []
    for img in sorted(path.iterdir()):
        print(str(img)[-7:], end=" ")
        # all_image, class_num = 0, 0
        img = Image.open(img)
        img = data_transform(img)
        img = torch.unsqueeze(img, dim=0)

        json_path = './class_indices.json'
        assert os.path.exists(json_path), "file: '{}' dose not exist.".format(json_path)

        json_file = open(json_path, "r")
        class_indict = json.load(json_file)

        with torch.no_grad():
            predict_cla = [0, 0, 0, 0, 0]
            output = torch.tensor([0., 0., 0., 0.]).to(device)
            for model in model_list:
                output += torch.squeeze(model(img.to(device)))
            output = output.cpu()
            predict = torch.softmax(output.cpu(), dim=0)
            predict_cla[torch.argmax(predict).numpy() + 1] += 1
            predict_cla = torch.argmax(torch.tensor(predict_cla))
            predict_cla = int(predict_cla) - 1
            res = "{:.3},{:.3},{:.3},{:.3}".format(output[0].numpy(), output[1].numpy(), output[2].numpy(), output[3].numpy())

        res_list.append(res)
        # print(res_list, len(res_list))
        print_res = "model class: {}  prob: {:.3}".format(class_indict[str(predict_cla)],
                                                     predict[predict_cla].numpy())
        print(print_res)

    for i in range(1, 318):
        sheet.cell(row=i, column=1, value=res_list[i-1])

    wb.save('test.xlsx')


if __name__ == '__main__':
    main()
